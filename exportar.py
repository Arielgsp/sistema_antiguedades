"""
exportar.py - Genera reportes en Excel (.xlsx):
  1) Listado de ascensos de un año determinado.
  2) Backup completo en CSV de todas las tablas (respaldo legible fuera de SQLite).

Uso:
    python exportar.py ascensos 2026
    python exportar.py backup_csv
"""
import sys
import csv
from pathlib import Path
from datetime import datetime

from db import get_connection, BASE_DIR
import operaciones as ops

EXPORT_DIR = BASE_DIR / "exports"
EXPORT_DIR.mkdir(exist_ok=True)


def exportar_ascensos_excel(anio: int, solo_1421: bool = True, fecha_corte=None):
    # Import diferido a propósito: así "backup_csv" (el respaldo pensado para
    # funcionar SIN depender de nada especial) sigue andando aunque openpyxl
    # no esté instalado o falle. Sólo hace falta para generar .xlsx.
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    resultados = ops.listar_ascensos_anio(anio, solo_1421=solo_1421, fecha_corte=fecha_corte)

    wb = Workbook()
    ws = wb.active
    universo = "Decreto 1421/02" if solo_1421 else "Todo el universo"
    ws.title = f"Ascensos {anio}"[:31]

    encabezados = ["N° Documento", "Apellido y Nombre", "Grado anterior", "Grado nuevo",
                   "Grados que suma", "Antigüedad computable", "Antigüedad (años)",
                   "Fecha efectiva del ascenso"]
    ws.append([f"Universo: {universo}"])
    ws.append(encabezados)
    for cell in ws[2]:
        cell.font = Font(bold=True, name="Arial", color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")

    for r in resultados:
        ws.append([
            r["n_doc"], r["apellido_nombre"], r["grados_anio_anterior"], r["grados_acumulados"],
            r["grados_nuevos"], r["antiguedad_computable_texto"], r["antiguedad_computable_anios"],
            r["fecha_efectiva_ascenso"],
        ])

    for col, ancho in zip("ABCDEFGH", [14, 38, 14, 12, 14, 26, 16, 22]):
        ws.column_dimensions[col].width = ancho
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.font = Font(name="Arial")

    destino = EXPORT_DIR / f"ascensos_{anio}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(destino)
    print(f"Generado: {destino}  ({len(resultados)} agentes ascienden)")
    return destino


def exportar_agentes_excel(solo_1421: bool = True):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    resultados = ops.listar_agentes_activos(solo_1421=solo_1421)

    wb = Workbook()
    ws = wb.active
    ws.title = "Agentes activos"

    encabezados = ["N° Documento", "Apellido y Nombre", "Nivel", "Grado", "Dependencia",
                   "Título", "Fecha de titulación", "Antigüedad ascenso 1421", "Antigüedad APN"]
    ws.append([f"Total agentes activos (Decreto 1421/02): {len(resultados)}"])
    ws.append(encabezados)
    for cell in ws[2]:
        cell.font = Font(bold=True, name="Arial", color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")

    for r in resultados:
        ws.append([
            r["n_doc"], r["apellido_nombre"], r["nivel_actual"], r["grado_actual"],
            r["dependencia_1421"], r["titulo"], r["fecha_titulacion"],
            r["antiguedad_1421_texto"], r["antiguedad_apn_texto"],
        ])

    for col, ancho in zip("ABCDEFGHI", [14, 38, 8, 8, 34, 16, 16, 24, 24]):
        ws.column_dimensions[col].width = ancho
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.font = Font(name="Arial")

    destino = EXPORT_DIR / f"agentes_activos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(destino)
    print(f"Generado: {destino}  ({len(resultados)} agentes)")
    return destino


def _categoria_ocupacional(n_doc: int, conn) -> tuple:
    """Clasifica a un agente en Profesional / Técnico / Administrativo según
    sus títulos activos (criterio definido por Ariel, agosto 2026):
      - Profesional: tiene título universitario (id_niv empieza con 'U').
      - Técnico: no es profesional, pero tiene título terciario
        (id_niv empieza con 'T').
      - Administrativo: no tiene ninguno de los dos (puede tener título
        secundario, no poseer título, u otra situación).
    Se prioriza Profesional sobre Técnico si tuviera ambos. Devuelve
    (categoria, id_niv_usado) -- el segundo valor es None para Administrativo."""
    filas = conn.execute(
        "SELECT id_niv FROM titulos WHERE n_doc=? AND activo=1 AND titulo != 'zz'", (n_doc,)
    ).fetchall()
    ids = [r["id_niv"] for r in filas if r["id_niv"]]
    for i in ids:
        if i.startswith("U"):
            return "Profesional", i
    for i in ids:
        if i.startswith("T"):
            return "Técnico", i
    return "Administrativo", None


def exportar_agrupamiento_excel(solo_1421: bool = True):
    """Informe de cantidad de agentes por categoría ocupacional (Profesional /
    Técnico / Administrativo, según título) cruzado por nivel (A a E) --
    pensado para responder pedidos de documentación externa sobre
    agrupamiento/escalafón. Los agentes sin nivel cargado quedan aparte, en
    una lista "a revisar", en vez de forzarlos en el cuadro."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    conn = get_connection()
    try:
        query = "SELECT n_doc, apellido_nombre, nivel_actual FROM agentes WHERE activo=1"
        if solo_1421:
            query += " AND cuenta_1421=1"
        query += " ORDER BY apellido_nombre"
        agentes = conn.execute(query).fetchall()

        niveles = ["A", "B", "C", "D", "E"]
        categorias = ["Profesional", "Técnico", "Administrativo"]
        matriz = {c: {n: 0 for n in niveles} for c in categorias}
        detalle = []
        a_revisar = []

        for a in agentes:
            categoria, id_niv = _categoria_ocupacional(a["n_doc"], conn)
            nivel = a["nivel_actual"]
            if nivel in niveles:
                matriz[categoria][nivel] += 1
            else:
                a_revisar.append((a["n_doc"], a["apellido_nombre"], categoria))
            detalle.append((a["n_doc"], a["apellido_nombre"], nivel or "(sin nivel)", categoria, id_niv or "-"))
    finally:
        conn.close()

    negrita_blanca = Font(bold=True, name="Arial", color="FFFFFF")
    relleno_azul = PatternFill("solid", fgColor="1F4E78")
    centrado = Alignment(horizontal="center")
    borde = Border(*(Side(style="thin"),) * 4)

    wb = Workbook()

    # --- Hoja 1: Resumen (cuadro categoría x nivel) ---
    ws = wb.active
    ws.title = "Resumen"
    universo = "agentes vigentes del Decreto 1421/02" if solo_1421 else "todos los agentes activos"
    ws.append([f"Cantidad de agentes por agrupamiento/categoría ocupacional y nivel"
               f" -- {universo} al {datetime.now().strftime('%d/%m/%Y')}"])
    ws.merge_cells("A1:G1")
    ws.append([])
    encabezados = ["Categoría ocupacional"] + niveles + ["Total"]
    ws.append(encabezados)
    for cell in ws[3]:
        cell.font = negrita_blanca
        cell.fill = relleno_azul
        cell.alignment = centrado
        cell.border = borde

    total_col = {n: 0 for n in niveles}
    for c in categorias:
        fila = [c] + [matriz[c][n] for n in niveles]
        fila.append(sum(matriz[c][n] for n in niveles))
        ws.append(fila)
        for n in niveles:
            total_col[n] += matriz[c][n]

    total_general = sum(total_col.values())
    ws.append(["Total"] + [total_col[n] for n in niveles] + [total_general])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, name="Arial")
        cell.border = borde

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row:
            if not cell.font.bold:
                cell.font = Font(name="Arial")
            cell.border = borde
            if cell.column_letter != "A":
                cell.alignment = centrado

    ws.append([])
    ws.append(["Criterio utilizado:"])
    ws.append(["Profesional: posee título universitario cargado en el sistema."])
    ws.append(["Técnico: no es profesional, pero posee título terciario cargado en el sistema."])
    ws.append(["Administrativo: no posee título universitario ni terciario cargado (puede tener"
               " título secundario, no poseer título, u otra situación)."])
    for r in range(ws.max_row - 3, ws.max_row + 1):
        ws.cell(row=r, column=1).font = Font(name="Arial", italic=True, size=9)

    if a_revisar:
        ws.append([])
        ws.append([f"A revisar -- {len(a_revisar)} agente(s) sin nivel cargado (no se pudieron"
                   " ubicar en el cuadro anterior):"])
        ws.cell(row=ws.max_row, column=1).font = Font(name="Arial", bold=True, size=9)
        for n_doc, nombre, categoria in a_revisar:
            ws.append([f"  {n_doc} - {nombre} - categoría estimada: {categoria}"])
            ws.cell(row=ws.max_row, column=1).font = Font(name="Arial", size=9)

    for col, ancho in zip("ABCDEFG", [22, 10, 10, 10, 10, 10, 10]):
        ws.column_dimensions[col].width = ancho

    # --- Hoja 2: Detalle por agente (respaldo/auditoría del cuadro) ---
    ws2 = wb.create_sheet("Detalle por agente")
    ws2.append(["N° Documento", "Apellido y Nombre", "Nivel", "Categoría ocupacional", "Título (id_niv)"])
    for cell in ws2[1]:
        cell.font = negrita_blanca
        cell.fill = relleno_azul
        cell.alignment = centrado
    for fila in detalle:
        ws2.append(list(fila))
    for col, ancho in zip("ABCDE", [14, 38, 10, 22, 16]):
        ws2.column_dimensions[col].width = ancho
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")

    destino = EXPORT_DIR / f"agrupamiento_ocupacional_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(destino)
    print(f"Generado: {destino}  ({len(agentes)} agentes, {len(a_revisar)} a revisar)")
    return destino


def exportar_backup_csv():
    """Exporta TODAS las tablas a CSV plano, como respaldo adicional legible
    (independiente de SQLite) en caso de necesitar recuperar datos manualmente."""
    conn = get_connection()
    tablas = ["agentes", "periodos_antiguedad", "config_agente", "titulos",
              "calculos_ascenso", "auditoria", "metadata"]
    carpeta = EXPORT_DIR / f"backup_csv_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    carpeta.mkdir(exist_ok=True)
    for t in tablas:
        rows = conn.execute(f"SELECT * FROM {t}").fetchall()
        if not rows:
            continue
        destino = carpeta / f"{t}.csv"
        with open(destino, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(rows[0].keys())
            for r in rows:
                writer.writerow(list(r))
        print(f"  {t}.csv ({len(rows)} filas)")
    conn.close()
    print(f"Backup CSV completo en: {carpeta}")
    return carpeta


def ultimo_backup_csv():
    """Fecha (datetime) del backup_csv más reciente que exista en /exports,
    o None si todavía no se generó ninguno."""
    carpetas = sorted(EXPORT_DIR.glob("backup_csv_*"))
    if not carpetas:
        return None
    ultima = carpetas[-1].name  # "backup_csv_AAAAMMDD_HHMMSS" ordena bien como texto
    try:
        return datetime.strptime(ultima, "backup_csv_%Y%m%d_%H%M%S")
    except ValueError:
        return None


def backup_csv_si_corresponde(dias: int = 7):
    """Genera un backup_csv automático si no hay ninguno o si el último ya
    tiene más de `dias` días. Pensado para llamarse solo, en cada arranque
    del programa (gui.py / cli.py) -- así no depende de que alguien se
    acuerde de correrlo a mano, ni de configurar un programador de tareas
    del sistema operativo (que además sería distinto en Windows y Mac).
    Devuelve la carpeta generada, o None si no hacía falta generar nada."""
    ultimo = ultimo_backup_csv()
    if ultimo and (datetime.now() - ultimo).days < dias:
        return None
    return exportar_backup_csv()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python exportar.py ascensos <anio>  |  python exportar.py backup_csv"
              "  |  python exportar.py agrupamiento")
        sys.exit(1)
    cmd = sys.argv[1]
    if cmd == "ascensos":
        anio = int(sys.argv[2])
        solo_1421 = "--todos" not in sys.argv
        exportar_ascensos_excel(anio, solo_1421=solo_1421)
    elif cmd == "backup_csv":
        exportar_backup_csv()
    elif cmd == "agrupamiento":
        solo_1421 = "--todos" not in sys.argv
        exportar_agrupamiento_excel(solo_1421=solo_1421)
    else:
        print("Comando desconocido.")
